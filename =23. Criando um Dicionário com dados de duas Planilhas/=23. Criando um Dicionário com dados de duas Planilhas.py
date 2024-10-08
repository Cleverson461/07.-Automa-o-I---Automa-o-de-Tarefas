from openpyxl import Workbook, load_workbook
from openpyxl.chart import BarChart, Reference, Series

dict_ano = {}

# 1 - Importando Despesas
arquivo1 = load_workbook(filename='files/despesas.xlsx')
ws1 = arquivo1['despesas']
max_linhas = ws1.max_row

for i in range(2, max_linhas+1):
    dict_ano[ws1['A%s' %i].value] = {'despesa':ws1['B%s' %i].value, 'receita':0}
    

# 2 - Importando Receita
arquivo2 = load_workbook(filename='files/receitas.xlsx')
ws2 = arquivo2['receitas']
max_linhas = ws2.max_row

for i in range(2, max_linhas+1):
    dict_ano[ws2['A%s' %i].value]['receita'] = ws2['B%s' %i].value
    
print(dict_ano)

# 3 - Criando a planilha
wb = Workbook()
ws = wb.active

ws['A1'] = 'Ano'
ws['B1'] = 'Despesas'
ws['C1'] = 'Receitas'

i = 2
for key, value in dict_ano.items():
    ws['A%s'%i] = key
    ws['B%s'%i] = value['despesa']
    ws['C%s'%i] = value['receita']
    i += 1
    
wb.save(filename='files/demonstrativo.xlsx')