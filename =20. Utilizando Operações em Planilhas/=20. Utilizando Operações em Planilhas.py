from openpyxl import load_workbook
from openpyxl.drawing.image import Image

wb = load_workbook(filename='files/gastos.xlsx')
planilha = wb['Sheet']

valor_total = 0
# 1 - Somando os valores 
for i in range(2, 10):
    valor = int(planilha['B%s' %i].value)
    valor_total += valor 

# print(valor_total)
planilha['B10'] = valor_total

# wb.save(filename='files/gastos.xlsx')

# 2 - Mesclar Células
planilha['A11'] = 'Teste'
planilha.merge_cells('A11:C11')
planilha.unmerge_cells('A11:C11')
# wb.save(filename='files/gastos.xlsx')

# 3 - Inserindo imagens
img = Image('bb_preco.png')
planilha.add_image(img, 'A12')
# wb.save(filename='files/gastos.xlsx')

# 4 - Deletando Células
planilha.delete_rows(1)
planilha.delete_cols(3)
wb.save(filename='files/gastos.xlsx')
